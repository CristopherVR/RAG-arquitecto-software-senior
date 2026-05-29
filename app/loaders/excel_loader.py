import openpyxl
from app.models.document import DocumentChunk


class ExcelLoader:

    @staticmethod
    def load_excel(file_path):

        chunks = []

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            for sheet in workbook.sheetnames:
                ws = workbook[sheet]

                headers = []
                for cell in ws[1]:
                    headers.append(str(cell.value) if cell.value else "")

                for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    values = [str(value) if value is not None else "" for value in row]

                    content = f"Diccionario de datos en hoja {sheet}, fila {row_number}:\n"

                    for header, value in zip(headers, values):
                        if header or value:
                            content += f"{header}: {value}\n"

                    chunks.append(DocumentChunk(
                        content=content,
                        metadata={
                            "source": file_path,
                            "file": file_path,
                            "sheet": sheet,
                            "row": row_number,
                            "type": "excel_dictionary"
                        }
                    ))

        except Exception as e:
            print(f"❌ Error leyendo Excel: {file_path}")
            print(e)

        return chunks