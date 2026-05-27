import os
import openpyxl
from app.models.document import DocumentChunk


class ExcelParser:

    @staticmethod
    def parse_file(file_path):
        """
        Parsea un archivo .xlsx que funciona como diccionario de datos.
        Retorna una lista de DocumentChunk:
          - Un chunk por cada tabla/entidad encontrada
          - Un chunk por cada campo individual (para trazabilidad precisa)
          - Un chunk de vista general por hoja
        """

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            print(f"Error abriendo {file_path}: {e}")
            return []

        file_name = os.path.basename(file_path)
        chunks = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            rows = []
            for row in ws.iter_rows(values_only=True):
                # Ignorar filas completamente vacías
                if any(cell is not None and str(cell).strip() != "" for cell in row):
                    rows.append([str(cell).strip() if cell is not None else "" for cell in row])

            if not rows:
                continue

            # La primera fila es el encabezado
            headers = rows[0]
            data_rows = rows[1:]

            if not data_rows:
                continue

            # Detectar columnas clave del diccionario de datos
            col_map = ExcelParser._map_columns(headers)

            # -------------------------
            # CHUNK DE VISTA GENERAL
            # -------------------------
            field_names = []
            for row in data_rows:
                name_col = col_map.get("name")
                if name_col is not None and name_col < len(row):
                    val = row[name_col]
                    if val:
                        field_names.append(val)

            overview_content = (
                f"Diccionario de datos - Hoja: {sheet_name}\n"
                f"Archivo: {file_name}\n"
                f"Columnas del encabezado: {', '.join(headers)}\n"
                f"Total de campos/registros: {len(data_rows)}\n"
            )

            if field_names:
                overview_content += f"\nCampos definidos:\n"
                overview_content += "\n".join(f"  - {f}" for f in field_names)

            chunks.append(DocumentChunk(
                content=overview_content,
                metadata={
                    "entity_type": "data_dictionary_overview",
                    "name": sheet_name,
                    "file_name": file_name,
                    "path": file_path,
                    "sheet": sheet_name,
                    "source": "excel",
                    "line": 1
                }
            ))

            # -------------------------
            # UN CHUNK POR CAMPO
            # (clave para trazabilidad: caso de uso 1 de la rúbrica)
            # -------------------------
            for row_index, row in enumerate(data_rows, start=2):

                field_content_parts = [
                    f"Campo del diccionario de datos",
                    f"Hoja: {sheet_name}",
                    f"Archivo: {file_name}",
                    f"Fila: {row_index}",
                ]

                field_name = ""

                for col_index, header in enumerate(headers):
                    if col_index < len(row) and row[col_index]:
                        field_content_parts.append(f"{header}: {row[col_index]}")

                        # Guardar el nombre del campo para el metadata
                        if col_index == col_map.get("name"):
                            field_name = row[col_index]

                if not field_name:
                    # Si no se detectó columna de nombre, usar el primer valor no vacío
                    field_name = next((v for v in row if v), f"campo_fila_{row_index}")

                chunks.append(DocumentChunk(
                    content="\n".join(field_content_parts),
                    metadata={
                        "entity_type": "data_field",
                        "name": field_name,
                        "file_name": file_name,
                        "path": file_path,
                        "sheet": sheet_name,
                        "source": "excel",
                        "line": row_index
                    }
                ))

            # -------------------------
            # CHUNK POR TABLA/ENTIDAD
            # (agrupa todos los campos de una misma tabla)
            # -------------------------
            table_col = col_map.get("table")

            if table_col is not None:
                tables = {}

                for row in data_rows:
                    if table_col < len(row) and row[table_col]:
                        table_name = row[table_col]
                        if table_name not in tables:
                            tables[table_name] = []
                        tables[table_name].append(row)

                for table_name, table_rows in tables.items():
                    field_lines = []

                    for row in table_rows:
                        parts = []
                        for col_index, header in enumerate(headers):
                            if col_index < len(row) and row[col_index] and col_index != table_col:
                                parts.append(f"{header}: {row[col_index]}")
                        field_lines.append("  " + " | ".join(parts))

                    table_content = (
                        f"Tabla del diccionario de datos: {table_name}\n"
                        f"Hoja: {sheet_name}\n"
                        f"Archivo: {file_name}\n"
                        f"Total de campos: {len(table_rows)}\n\n"
                        f"Campos:\n" + "\n".join(field_lines)
                    )

                    chunks.append(DocumentChunk(
                        content=table_content,
                        metadata={
                            "entity_type": "data_table",
                            "name": table_name,
                            "file_name": file_name,
                            "path": file_path,
                            "sheet": sheet_name,
                            "source": "excel",
                            "line": 1
                        }
                    ))

        wb.close()
        return chunks

    # -------------------------------------------------------------------------
    # Helpers privados
    # -------------------------------------------------------------------------

    @staticmethod
    def _map_columns(headers):
        """
        Detecta qué columna corresponde a cada rol clave del diccionario
        de datos (nombre del campo, tipo, tabla, descripción, etc.)
        buscando palabras clave en los encabezados.
        """

        col_map = {}
        headers_lower = [h.lower() for h in headers]

        NAME_KEYS    = ["campo", "field", "nombre", "name", "columna", "column", "atributo"]
        TABLE_KEYS   = ["tabla", "table", "entidad", "entity", "modelo", "model"]
        TYPE_KEYS    = ["tipo", "type", "datatype", "data type", "tipo de dato"]
        DESC_KEYS    = ["descripcion", "description", "detalle", "detail", "comentario"]
        PK_KEYS      = ["pk", "primary", "clave primaria", "llave"]
        FK_KEYS      = ["fk", "foreign", "clave foranea", "referencia"]
        NULL_KEYS    = ["null", "nulo", "nullable", "requerido", "required"]

        def find_col(keys):
            for i, h in enumerate(headers_lower):
                if any(k in h for k in keys):
                    return i
            return None

        col_map["name"]  = find_col(NAME_KEYS)
        col_map["table"] = find_col(TABLE_KEYS)
        col_map["type"]  = find_col(TYPE_KEYS)
        col_map["desc"]  = find_col(DESC_KEYS)
        col_map["pk"]    = find_col(PK_KEYS)
        col_map["fk"]    = find_col(FK_KEYS)
        col_map["null"]  = find_col(NULL_KEYS)

        return col_map